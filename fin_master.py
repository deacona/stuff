#!/usr/bin/python -tt
"""
Created on Mon Feb 06 16:49:37 2017

@author: adeacon
"""

import sys
import os
import re
import zipfile
from openpyxl import load_workbook
import subprocess
import pandas as pd
from datetime import datetime

# from pandas.tools.plotting import scatter_matrix
import numpy as np

# import matplotlib.pyplot as plt
from sklearn import model_selection
from sklearn.metrics import classification_report
from sklearn.metrics import confusion_matrix
from sklearn.metrics import accuracy_score

# from sklearn.linear_model import LogisticRegression
# from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier

# from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
# from sklearn.naive_bayes import GaussianNB
# from sklearn.svm import SVC
import config

pd.set_option("display.max_columns", 500)
pd.set_option("display.expand_frame_repr", False)

source_path = config.SOURCE_DIR + "/snt"
account_dict = config.FIN_ACCOUNTS
finances_dict = config.FIN_DOCS
group_dict = config.FIN_GROUPS


def get_date_dim(start_date):
    df = pd.read_excel(config.MASTER_FILES["date_dim"])
    df = df[df["date key"] >= start_date]
    df = df[
        [
            "full date",
            "weekday flag",
            "week begin date",
            "quarter",
            "last day in month flag",
            "same day year ago",
        ]
    ]
    df.rename(columns={"full date": "Date"}, inplace=True)
    df.columns = df.columns.str.title().str.replace(" ", "_")

    return df


def date_expansion(df):
    df["Date"] = pd.to_datetime(df["Date"], format="%d/%m/%Y")
    df["YearMonth"] = df["Date"].apply(lambda x: x.strftime("%Y%m"))
    df["Year"] = df["Date"].apply(lambda x: x.strftime("%Y"))
    df["Month"] = df["Date"].apply(lambda x: x.strftime("%m"))
    df["Month_Abbrev"] = df["Date"].apply(lambda x: x.strftime("%b"))
    df["Day_Of_Month"] = df["Date"].apply(lambda x: x.strftime("%d"))
    df["Day_Of_Week"] = df["Date"].apply(lambda x: x.strftime("%w"))
    df["Day_Abrrev"] = df["Date"].apply(lambda x: x.strftime("%a"))
    df["DateKey"] = df["Date"].apply(lambda x: x.strftime("%Y%m%d"))

    df.DateKey = pd.to_numeric(df.DateKey, errors="coerce").fillna(0).astype(np.int64)

    datedim = get_date_dim(df["DateKey"].min())
    df = pd.merge(df, datedim, on="Date", how="inner")

    return df


def ensure_dir(f):
    d = os.path.dirname(f)
    if not os.path.exists(d):
        os.makedirs(d)


def save_master(dframe):
    dframe.to_csv(config.MASTER_FILES["fin_master"], encoding="utf-8", sep="|")
    print("\n... saved to " + config.MASTER_FILES["fin_master"])


def get_master():
    return pd.read_csv(config.MASTER_FILES["fin_master"], encoding="utf8", sep="|")


def get_type(x):
    return (
        str(x)
        .split(" TO ")[0]
        .split(" FROM ")[0]
        .split(" AT ")[0]
        .split(" REF ")[0]
        .split(" REF.")[0]
    )


def clean_desc(x):
    for start in [" TO ", " FROM ", " AT ", " REF ", " REF."]:
        if str(x).count(start) > 0:
            x = x.split(start)[1]
    return str(x).split(",")[0]


def func_debit_credit(x):
    if x < 0:
        return "Debit", (-1 * x)
    else:
        return "Credit", x


def func_account(filename):
    for key, value in account_dict.items():
        if filename.startswith(key):
            return value[1]


def get_legacy_data():
    df = pd.read_csv(
        config.MASTER_FILES["fin_legacy"],
        header=0,
        names=[
            "Date",
            "Description",
            "Credit_amount",
            "Debit_amount",
            "BalanceText",
            "Year",
            "Month",
            "YearMonth",
            "Debit_Credit_Flag",
            "Amount",
            "Balance",
            "Account",
            "Type",
            "Category",
        ],
    )
    df = df.dropna(subset=["Date", "Description"])

    df["Amount"] = pd.to_numeric(df["Amount"].str.replace(",", ""), errors="coerce")
    df["Balance"] = pd.to_numeric(df["Balance"].str.replace(",", ""), errors="coerce")
    df["Filename"] = config.MASTER_FILES["fin_legacy"].split(os.sep)[-1]

    df = date_expansion(df)

    df["Transaction"] = df["Amount"]
    df["Debit_Credit_Flag"], df["Amount"] = list(
        zip(*df["Transaction"].map(func_debit_credit))
    )
    df["Account_Holder"] = "Joint"  # df['Filename'].map(func_account)
    df["Type"] = df["Description"].map(get_type)
    df["Description_Clean"] = df["Description"].map(clean_desc)
    df["Description_Item"] = (
        df["Description_Clean"] + " // " + df["Type"] + " // " + df["Debit_Credit_Flag"]
    )

    return df


def get_categories():
    df = get_legacy_data()

    cat = pd.DataFrame(
        df.groupby("Description_Clean")["Category"].agg(
            lambda x: x.value_counts().index[0]
        )
    )
    cat = cat.reset_index()

    return cat


def nonblank_lines(f):
    for l in f:
        line = l.rstrip()
        if line:
            yield line


# def clean_text(row):
#     return [r.decode("unicode_escape").encode("ascii", "ignore") for r in row]


def get_statements(filename):
    row_dict = {}
    full_data = []
    i = 0
    with open(os.path.join(source_path, filename)) as in_file:
        for line in nonblank_lines(in_file):
            name, var = line.partition(":")[::2]
            if name != "From":
                if name == "Date":
                    i = 0
                else:
                    i = i + 1
                row_dict[name.strip()] = var.strip()
                if i == 3:
                    full_data.append(row_dict.copy())

    df = pd.DataFrame(full_data)

    df = df.drop_duplicates()
    df["Filename"] = filename
    # df[["Account", "Amount", "Balance", "Date", "Description"]] = df[
    #     ["Account", "Amount", "Balance", "Date", "Description"]
    # ].apply(clean_text)
    df["Amount"] = pd.to_numeric(df["Amount"].str.replace("GBP", ""), errors="coerce")
    df["Balance"] = pd.to_numeric(df["Balance"].str.replace("GBP", ""), errors="coerce")

    df = date_expansion(df)

    df["Transaction"] = df["Amount"]
    df["Debit_Credit_Flag"], df["Amount"] = list(
        zip(*df["Transaction"].map(func_debit_credit))
    )
    df["Account_Holder"] = df["Filename"].map(func_account)
    df["Type"] = df["Description"].map(get_type)
    df["Description_Clean"] = df["Description"].map(clean_desc)
    df["Description_Item"] = (
        df["Description_Clean"] + " // " + df["Type"] + " // " + df["Debit_Credit_Flag"]
    )

    # print df.head(1)
    return df


def data_validator(df):
    df["checks"] = df["Transaction"].shift(1) + df["Balance"].diff()
    # checks = df[["Filename","Transaction","Balance","checks"]][(df["checks"] > 1) & (df["Filename"].shift(1) == df["Filename"]) & (df["Filename"] <> "fin_legacy.csv")]
    # checks = df[["Account_Holder","Filename","Date","Transaction","Balance","checks"]][(df["checks"] > 1) & (df["Filename"].shift(1) == df["Filename"])]
    checks = df[
        ["Account_Holder", "Filename", "Date", "Transaction", "Balance", "checks"]
    ][
        (df["checks"] >= 1)
        & (df["Account_Holder"].shift(1) == df["Account_Holder"])
        & (df["Filename"] != "fin_legacy.csv")
    ]

    if not checks.empty:
        limit = 5
        print("First {0} differences...".format(str(limit)))
        print(checks.head(limit))
        print("...")
        sys.exit("Significant diffs found in Balance vs Transactions")


def get_account_data():
    ### TODO handle other account formats

    previous_account = ""
    for root, dirs, files in os.walk(source_path, topdown=False):
        for i, filename in enumerate(sorted(files)):
            print(i, filename)
            df_new = get_statements(filename)
            account = filename[10:24]
            if i == 0:
                df = df_new[0:0]
                df_end = 0  #'000000'
            elif account != previous_account:
                df_end = 0  #'000000'
            else:
                df_end = df[df["Filename"].str.contains(account)]["DateKey"].max()

            df = pd.concat([df_new[df_new["DateKey"] > df_end], df])
            previous_account = account

    ### TODO improve classification process
    cat = get_categories()
    df = pd.merge(df, cat, on="Description_Clean", how="left")

    old = get_legacy_data()
    old_end = old["DateKey"].max()
    df = pd.concat([df[df["DateKey"] > old_end], old], sort=True)

    df["Category_Group"] = df["Category"].map(group_dict)
    df["Category_Group"].fillna(config.FIN_GROUPS["Unknown"], inplace=True)

    ### TODO populate "Bank","Merchant","Place"
    for dummy in ["Lineno", "Bank", "Merchant", "Place"]:
        df[dummy] = np.nan

    # Filename|Lineno|Account_Holder|Bank|Date|Year|Month|Day_Of_Month|Day_Of_Week|YearMonth|DateKey|Day_Abrrev|Weekday_Flag|Week_Begin_Date|Month_Abbrev|Quarter|Last_Day_In_Month_Flag|Same_Day_Year_Ago|Type|Description|Description_Clean|Description_Item|Category|Merchant|Place|Category_Group|Debit_Credit_Flag|Transaction|Amount|Balance
    cols = [
        "Filename",
        "Lineno",
        "Account_Holder",
        "Bank",
        "Date",
        "Year",
        "Month",
        "Day_Of_Month",
        "Day_Of_Week",
        "YearMonth",
        "DateKey",
        "Day_Abrrev",
        "Weekday_Flag",
        "Week_Begin_Date",
        "Month_Abbrev",
        "Quarter",
        "Last_Day_In_Month_Flag",
        "Same_Day_Year_Ago",
        "Type",
        "Description",
        "Description_Clean",
        "Description_Item",
        "Category",
        "Merchant",
        "Place",
        "Category_Group",
        "Debit_Credit_Flag",
        "Transaction",
        "Amount",
        "Balance",
    ]
    df = df[cols]

    data_validator(df)
    print(df.groupby("Filename").agg({"DateKey": [min, max, "count"]}))
    # df.info()
    # print df.describe(include="all")
    save_master(df)

    return df


def get_feature_vector():
    df = get_legacy_data()
    # Transaction
    # Type --> 0/1 flags
    # Description_Clean --> levenshtein distance (from most frequest?)

    print(df["Type"].value_counts().head(7))

    # df.info()
    # print df.describe(include="all")


def evaluate_algorithms():
    dataset = get_legacy_data()
    dataset["Category_Group"] = dataset["Category"].map(group_dict)
    dataset["Category_Group"].fillna(config.FIN_GROUPS["Unknown"], inplace=True)

    # array = dataset[["Transaction", "Amount", "Balance", "DateKey", "Category_Group"]].values
    array = dataset[["Transaction", "Amount", "Category_Group"]].values
    # print array[:5]
    print(array.shape)

    # pick columns to use
    X = array[:, 0:1]
    # print X
    Y = array[:, 2]
    # print Y
    validation_size = 0.20
    seed = 7
    X_train, X_validation, Y_train, Y_validation = model_selection.train_test_split(
        X, Y, test_size=validation_size, random_state=seed
    )

    #    print X_train[:5]
    #    print X_validation[:5]
    #    print Y_train[:5]
    #    print Y_validation[:5]

    #    # Test options and evaluation metric
    #    seed = 7
    #    scoring = 'accuracy'
    #
    #    # Spot Check Algorithms
    #    models = []
    #    models.append(('LR', LogisticRegression()))
    #    models.append(('LDA', LinearDiscriminantAnalysis()))
    #    models.append(('KNN', KNeighborsClassifier()))
    #    models.append(('CART', DecisionTreeClassifier()))
    #    models.append(('NB', GaussianNB()))
    #    models.append(('SVM', SVC()))
    #    # evaluate each model in turn
    #    results = []
    #    names = []
    #    for name, model in models:
    #        kfold = model_selection.KFold(n_splits=10, random_state=seed)
    #        cv_results = model_selection.cross_val_score(model, X_train, Y_train, cv=kfold, scoring=scoring)
    #        results.append(cv_results)
    #        names.append(name)
    #        msg = "%s: %f (%f)" % (name, cv_results.mean(), cv_results.std())
    #        print(msg)
    #
    #    # Compare Algorithms
    #    fig = plt.figure()
    #    fig.suptitle('Algorithm Comparison')
    #    ax = fig.add_subplot(111)
    #    plt.boxplot(results)
    #    ax.set_xticklabels(names)
    #    plt.show()

    # Make predictions on validation dataset
    print("\nPrediction using K Nearest Neighbours algorithm...")
    knn = KNeighborsClassifier()
    knn.fit(X_train, Y_train)
    predictions = knn.predict(X_validation)
    print((accuracy_score(Y_validation, predictions)))
    print((confusion_matrix(Y_validation, predictions)))
    print((classification_report(Y_validation, predictions)))
    # print predictions

    results = knn.predict(X)
    dataset["Predicted_Category_Group"] = results
    # dataset.info()
    # print dataset.describe(include="all")
    print(
        dataset[dataset["Predicted_Category_Group"] != dataset["Category_Group"]].head(
            1
        )
    )


#    print dataset.tail(5)

#    print "\nPrediction using Support Vector Clustering algorithm..."
#    svc = SVC()
#    svc.fit(X_train, Y_train)
#    predictions = svc.predict(X_validation)
#    print(accuracy_score(Y_validation, predictions))
#    print(confusion_matrix(Y_validation, predictions))
#    print(classification_report(Y_validation, predictions))


def get_new_files():
    for filename in os.listdir(config.DOWNLOAD_DIR):
        # print filename
        if filename.startswith(tuple(account_dict.keys())):
            for prefix in list(account_dict.keys()):
                if filename.startswith(prefix):
                    outfile = (
                        prefix
                        + "_{0}.".format(
                            datetime.strftime(datetime.now(), "%Y%m%d%H%M")
                        )
                        + filename[-3:]
                    )  # str(filename.split(".")[-1:])[2:5]
                    # print outfile
                    print(
                        "Moving "
                        + os.path.join(config.DOWNLOAD_DIR, filename)
                        + " to "
                        + os.path.join(
                            config.SOURCE_DIR, account_dict[prefix][0], outfile
                        )
                    )
                    os.rename(
                        os.path.join(config.DOWNLOAD_DIR, filename),
                        os.path.join(
                            config.SOURCE_DIR, account_dict[prefix][0], outfile
                        ),
                    )
                    print("===> DONE")


def check_thing():
    thing = config.FIN_PASS
    # thing = input("Enter the thing:")
    if len(thing) == 0:
        sys.exit("thing cannot be empty")
    elif re.match("^[\w]+$", thing) is None:
        sys.exit("thing has invalid chars")
    else:
        #        sys.exit("all good")
        return thing.encode("utf-8")


def archive_files():
    thing = check_thing()

    zPath = "C:\Program Files\\7-Zip"
    zApp = "7z.exe"
    zAction = "a"
    zPass = "-p{0}".format(thing)
    zAnswer = "-y"
    zProg = os.path.join(zPath, zApp)

    for zipprefix, zipfolder in finances_dict.items():
        zipname = os.path.join(zipfolder, zipprefix + ".zip")
        filelist = []
        for filename in os.listdir(zipfolder):
            if filename.startswith(zipprefix) and not filename.endswith(".zip"):
                filelist.append(os.path.join(zipfolder, filename))

        cmd = [zApp, zAction, zipname, zPass, zAnswer] + filelist
        rc = subprocess.Popen(
            cmd, executable=zProg, stderr=subprocess.STDOUT, stdout=subprocess.PIPE
        )
        rc_out = rc.stdout.readlines()
        print(str("rc_out:{0}".format(rc_out)))

        for fileloc in filelist:
            print("fileloc:{0}".format(fileloc))
            os.remove(fileloc)


def extract_files():
    thing = check_thing()

    for zipprefix, zipfolder in finances_dict.items():
        if os.path.exists(os.path.join(zipfolder, zipprefix + ".zip")):
            print(zipprefix, zipfolder)
            os.chdir(zipfolder)
            archiveZip = zipfile.ZipFile(zipprefix + ".zip")
            # archiveZip.extractall()
            archiveZip.extractall(pwd=thing)
            archiveZip.close()
            os.remove(os.path.join(zipfolder, zipprefix + ".zip"))


def update_fin_analysis(df):
    writer = pd.ExcelWriter(config.MASTER_FILES["fin_datasheet"])
    df.to_excel(writer, sheet_name="fin_master", index=False)
    writer.save()
    print("... fin_analysis saved")


def update_cash_calculator(df):
    ### TODO add cash_calculator
    pass


def update_budget_planner(df):
    inputs = (
        df[df["Date"] >= df["Same_Day_Year_Ago"].max()]
        .groupby("Category_Group")
        .sum()["Transaction"]
        .abs()
    )

    wb = load_workbook(config.MASTER_FILES["fin_budget_template"])

    sheet = wb.get_sheet_by_name("What Do You Earn")
    sheet.cell(row=20, column=5).value = inputs[
        "i01. Income From Employment / Self Employment"
    ]
    sheet.cell(row=21, column=5).value = inputs[
        "i02. Income From Savings & Investments"
    ]

    sheet = wb.get_sheet_by_name("What Do You Spend")
    sheet.cell(row=16, column=8).value = inputs["o01. In Your Home"]
    sheet.cell(row=44, column=8).value = inputs["o02. Insurance"]
    sheet.cell(row=62, column=8).value = inputs["o03. Eats, drinks & smokes"]
    sheet.cell(row=81, column=8).value = inputs["o04. Motoring & Public Transport"]
    sheet.cell(row=141, column=8).value = inputs["o07. Family"]
    sheet.cell(row=162, column=8).value = inputs["o08. Fun & Frolics"]
    sheet.cell(row=240, column=8).value = inputs["o12. Big One Offs"]
    sheet.cell(row=259, column=8).value = inputs["o13. Odds & Sods"]

    ### TODO add in Monthly Desired Spend

    ### TODO save with date range in name
    wb.save(config.MASTER_FILES["fin_budget_output"])
    print("... budget_planner saved")


def export_viz(df):

    ### TODO export graphs
    ### 1. Balance over time (Min/Max Balance, Joint, all months)
    # df.info()
    bal = (
        df[df["Account_Holder"] == "Joint"]
        .groupby("DateKey")
        .agg({"Balance": [min, max, np.mean, np.median]})
    )

    bal = bal.tail(200)
    bal.info()
    print(bal.describe(include="all"))
    print(bal.tail(10))

    # bal.plot(kind='line', title="Joint Account Balance", figsize=(30, 10))


#    plt.show()

#    dfplot.set_xlabel("Year")
#    dfplot.set_ylabel(ylabel)

# plt.savefig(config.ANALYSIS_DIR+'/'+dfplotfile)


### 2. Monthly P&L (Sum of Transaction, Joint, all months)
### 3. Monthly transactions by Category (Sum of Amount, 3/4 months)


def open_and_wait():
    # print "call Popen"
    # proc1 = subprocess.Popen(r"C:\Windows\System32\calc.exe", shell=True))
    # print config.MASTER_FILES["fin_datasheet"]
    file1 = config.MASTER_FILES["fin_datasheet"].replace("/", "//")
    # print file1
    proc1 = subprocess.Popen(["explorer", file1])
    # proc2 = subprocess.Popen(r"C:\Windows\System32\calc.exe", shell=True))
    # print config.MASTER_FILES["fin_budget_output"]
    file2 = config.MASTER_FILES["fin_budget_output"].replace("/", "//")
    # print file2
    proc2 = subprocess.Popen(["explorer", file2])
    # print "wait for close"
    exit_codes = [p.wait() for p in (proc1, proc2)]
    # print "all done"


def send_for_analysis():
    df = get_master()

    update_fin_analysis(df)

    update_cash_calculator(df)

    update_budget_planner(df)

    export_viz(df)

    open_and_wait()


def check_with_user(text):
    userInput = input((text + " (y/n): ").rjust(50))
    if userInput[:1].lower() == "y":
        return True
    else:
        return False


def main():
    ### test components
    # get_statements("BLAH")
    # get_legacy_data()
    # get_categories()
    # get_feature_vector()
    # evaluate_algorithms()
    # get_date_dim(20150101)
    # date_expansion(get_legacy_data())
    # func_account("BLAH")
    # data_validator(get_master())
    # export_viz(get_master())
    # open_and_wait()
    # print(check_with_user("TEST TEXT"))

    ### end 2 end
    if check_with_user("Step 1: EXTRACTING ARCHIVED FILES"):
        extract_files()
    if check_with_user("Step 2: GET NEW FILES"):
        get_new_files()
    if check_with_user("Step 3: GET ACCOUNT DATA"):
        get_account_data()
    if check_with_user("Step 4: SEND FOR ANALYSIS"):
        send_for_analysis()
    if check_with_user("Step 5: ARCHIVE FILES"):
        archive_files()


if __name__ == "__main__":
    main()
